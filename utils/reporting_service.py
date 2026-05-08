import json
import os
import re
import tempfile

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class AnalysisReportingService:
    ASPECT_NAMES = ['الطعام', 'الخدمة', 'السعر', 'الأجواء', 'النظافة', 'الموقع', 'التوصيل']

    @staticmethod
    def calculate_stats(results):
        return {
            'total_reviews': len(results),
            'sentiment_distribution': {
                'إيجابي': sum(1 for r in results if r.get('overall_sentiment') == 'إيجابي'),
                'سلبي': sum(1 for r in results if r.get('overall_sentiment') == 'سلبي'),
                'محايد': sum(1 for r in results if r.get('overall_sentiment', 'محايد') == 'محايد')
            }
        }

    @classmethod
    def generate_local_executive_summary(cls, results):
        if not results:
            return "## ملخص تنفيذي\n\nلا توجد بيانات كافية لتوليد الملخص."

        stats = cls.calculate_stats(results)
        total = stats['total_reviews']
        positive = stats['sentiment_distribution']['إيجابي']
        negative = stats['sentiment_distribution']['سلبي']
        neutral = stats['sentiment_distribution']['محايد']

        aspect_stats = cls._collect_aspect_stats(results)

        ranked_positive = sorted(
            aspect_stats.items(),
            key=lambda item: (item[1]['إيجابي'] - item[1]['سلبي'], item[1]['إيجابي']),
            reverse=True
        )
        ranked_negative = sorted(
            aspect_stats.items(),
            key=lambda item: (item[1]['سلبي'] - item[1]['إيجابي'], item[1]['سلبي']),
            reverse=True
        )

        top_positive = [name for name, values in ranked_positive if values['إيجابي'] > 0][:3]
        top_negative = [name for name, values in ranked_negative if values['سلبي'] > 0][:3]

        if positive >= negative and positive >= neutral:
            overall_tone = "الانطباع العام يميل إلى الإيجابية، ما يشير إلى رضا جيد لدى شريحة كبيرة من العملاء."
        elif negative >= positive and negative >= neutral:
            overall_tone = "الانطباع العام يميل إلى السلبية، ما يعني وجود نقاط تشغيلية تستحق المعالجة السريعة."
        else:
            overall_tone = "الانطباع العام متوازن إلى حد كبير، مع وجود فرص واضحة لتحويل التقييمات المحايدة إلى إيجابية."

        strengths_text = "، ".join(top_positive) if top_positive else "لا توجد جوانب إيجابية بارزة بشكل كافٍ في البيانات الحالية"
        weaknesses_text = "، ".join(top_negative) if top_negative else "لا توجد جوانب سلبية مهيمنة بشكل واضح في البيانات الحالية"

        recommendations = cls._build_recommendations(top_negative)

        summary_lines = [
            "## ملخص تنفيذي",
            f"- تم تحليل **{total}** تقييم.",
            f"- التوزيع العام للمشاعر: **{positive} إيجابي**، **{negative} سلبي**، **{neutral} محايد**.",
            "",
            "## قراءة عامة",
            overall_tone,
            "",
            "## نقاط القوة",
            f"- أبرز الجوانب الإيجابية: **{strengths_text}**.",
            "",
            "## الجوانب التي تحتاج تحسين",
            f"- الجوانب الأكثر حاجة للمعالجة: **{weaknesses_text}**.",
            "",
            "## توصيات تنفيذية",
        ]
        summary_lines.extend([f"- {item}" for item in recommendations[:3]])
        summary_lines.extend([
            "",
            "## مقترحات للمرحلة القادمة",
            "- الاستمرار في تتبع تغيّر المشاعر عبر الملفات الجديدة لقياس أثر التحسينات.",
            "- استخدام الجوانب الأعلى سلبية كأولوية تشغيلية في خطط التطوير القادمة."
        ])
        return "\n".join(summary_lines)

    @classmethod
    def build_export_dataframe(cls, results):
        rows = []
        for item in results:
            row = {
                'النص الأصلي': item.get('original_text', ''),
                'النص المنظف': item.get('cleaned_text', ''),
                'المشاعر العامة': item.get('overall_sentiment', ''),
                'نسبة الثقة العامة': round(float(item.get('confidence', 0) or 0) * 100, 1),
                'الشرح التحليلي': item.get('logic_explanation', ''),
                'مزود التحليل': item.get('provider', ''),
            }

            aspects = item.get('aspects', []) or []
            aspect_map = {aspect.get('aspect'): aspect for aspect in aspects if aspect.get('aspect')}
            row['الجوانب المكتشفة'] = '، '.join(aspect_map.keys())

            for aspect_name in cls.ASPECT_NAMES:
                aspect_data = aspect_map.get(aspect_name, {})
                row[f'{aspect_name} - المشاعر'] = aspect_data.get('sentiment', '')
                confidence = aspect_data.get('confidence')
                row[f'{aspect_name} - الثقة %'] = round(float(confidence) * 100, 1) if confidence is not None else ''
                row[f'{aspect_name} - الشاهد'] = aspect_data.get('evidence', '') or ''
                implicit_value = aspect_data.get('implicit')
                row[f'{aspect_name} - النوع'] = 'ضمني' if implicit_value else ('صريح' if aspect_name in aspect_map else '')

            rows.append(row)

        return pd.DataFrame(rows)

    @staticmethod
    def strip_markdown_text(text):
        if not text:
            return ""
        text = re.sub(r'[`#*>\-]+', ' ', text)
        text = re.sub(r'\n{2,}', '\n', text)
        return text.strip()

    @classmethod
    def build_stats_dataframe(cls, results):
        stats = cls.calculate_stats(results)
        total = stats['total_reviews']
        positive = stats['sentiment_distribution']['إيجابي']
        negative = stats['sentiment_distribution']['سلبي']
        neutral = stats['sentiment_distribution']['محايد']
        aspect_stats = cls._collect_aspect_stats(results)

        rows = [{
            'المؤشر': 'إجمالي التقييمات',
            'القيمة': total,
            'النسبة %': 100 if total else 0
        }, {
            'المؤشر': 'إيجابي',
            'القيمة': positive,
            'النسبة %': round((positive / total) * 100, 1) if total else 0
        }, {
            'المؤشر': 'سلبي',
            'القيمة': negative,
            'النسبة %': round((negative / total) * 100, 1) if total else 0
        }, {
            'المؤشر': 'محايد',
            'القيمة': neutral,
            'النسبة %': round((neutral / total) * 100, 1) if total else 0
        }]

        for aspect, values in sorted(aspect_stats.items()):
            total_aspect = values['إيجابي'] + values['سلبي'] + values['محايد']
            rows.append({
                'المؤشر': f'جانب: {aspect}',
                'القيمة': total_aspect,
                'النسبة %': round((total_aspect / total) * 100, 1) if total else 0
            })

        return pd.DataFrame(rows), aspect_stats

    @staticmethod
    def style_excel_sheet(worksheet, sentiment_columns=None):
        header_fill = PatternFill(fill_type='solid', fgColor='D9EEE9')
        header_font = Font(bold=True, color='1A2E2B')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wrap_alignment = Alignment(vertical='top', wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        if sentiment_columns:
            positive_fill = PatternFill(fill_type='solid', fgColor='E8F5E9')
            negative_fill = PatternFill(fill_type='solid', fgColor='FDECEC')
            neutral_fill = PatternFill(fill_type='solid', fgColor='FFF6DB')

            for col_idx in sentiment_columns:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value == 'إيجابي':
                        cell.fill = positive_fill
                    elif cell.value == 'سلبي':
                        cell.fill = negative_fill
                    elif cell.value == 'محايد':
                        cell.fill = neutral_fill
                    cell.alignment = center_alignment

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column not in (1, 2, 5, 6) and not sentiment_columns:
                    cell.alignment = wrap_alignment

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 14), 45)

    @classmethod
    def export_results_to_excel(cls, results, summary_text, base_name):
        export_df = cls.build_export_dataframe(results)
        stats_df, aspect_stats = cls.build_stats_dataframe(results)
        summary_lines = [line.strip() for line in cls.strip_markdown_text(summary_text).splitlines() if line.strip()]
        summary_df = pd.DataFrame({'الملخص التنفيذي': summary_lines or ['لا توجد بيانات كافية لعرض الملخص التنفيذي.']})

        aspect_summary_rows = []
        for aspect, values in sorted(aspect_stats.items()):
            total_aspect = values['إيجابي'] + values['سلبي'] + values['محايد']
            aspect_summary_rows.append({
                'الجانب': aspect,
                'إيجابي': values['إيجابي'],
                'سلبي': values['سلبي'],
                'محايد': values['محايد'],
                'الإجمالي': total_aspect
            })
        aspect_summary_df = pd.DataFrame(aspect_summary_rows)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()

        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Analysis')
            summary_df.to_excel(writer, index=False, sheet_name='Executive Summary')
            stats_df.to_excel(writer, index=False, sheet_name='Management Stats', startrow=0)
            if not aspect_summary_df.empty:
                aspect_summary_df.to_excel(writer, index=False, sheet_name='Management Stats', startrow=len(stats_df) + 3)

            analysis_sheet = writer.sheets['Analysis']
            sentiment_columns = []
            for idx, cell in enumerate(analysis_sheet[1], start=1):
                if 'المشاعر' in str(cell.value or ''):
                    sentiment_columns.append(idx)
            cls.style_excel_sheet(analysis_sheet, sentiment_columns=sentiment_columns)

            summary_sheet = writer.sheets['Executive Summary']
            cls.style_excel_sheet(summary_sheet)
            for row in range(2, summary_sheet.max_row + 1):
                summary_sheet.cell(row=row, column=1).alignment = Alignment(vertical='top', wrap_text=True)
            summary_sheet.column_dimensions['A'].width = 120

            stats_sheet = writer.sheets['Management Stats']
            cls.style_excel_sheet(stats_sheet)
            for row in range(2, stats_sheet.max_row + 1):
                metric = stats_sheet.cell(row=row, column=1).value
                if metric in ('إيجابي',):
                    stats_sheet.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='E8F5E9')
                elif metric in ('سلبي',):
                    stats_sheet.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='FDECEC')
                elif metric in ('محايد',):
                    stats_sheet.cell(row=row, column=2).fill = PatternFill(fill_type='solid', fgColor='FFF6DB')

        return temp_file.name, f'analyzed_{base_name}.xlsx'

    @classmethod
    def load_summary_text(cls, summary_path, results):
        summary_text = cls.generate_local_executive_summary(results)
        if os.path.exists(summary_path):
            try:
                with open(summary_path, 'r', encoding='utf-8') as f:
                    summary_data = json.load(f)
                    summary_text = summary_data.get('executive_summary') or summary_text
            except Exception:
                pass
        return summary_text

    @classmethod
    def _collect_aspect_stats(cls, results):
        aspect_stats = {}
        for item in results:
            for aspect_data in item.get('aspects', []):
                aspect = aspect_data.get('aspect')
                sentiment = aspect_data.get('sentiment', 'محايد')
                if not aspect:
                    continue
                if aspect not in aspect_stats:
                    aspect_stats[aspect] = {'إيجابي': 0, 'سلبي': 0, 'محايد': 0}
                if sentiment in aspect_stats[aspect]:
                    aspect_stats[aspect][sentiment] += 1
        return aspect_stats

    @staticmethod
    def _build_recommendations(top_negative):
        recommendations = []
        if top_negative:
            for aspect in top_negative[:2]:
                if aspect == 'الخدمة':
                    recommendations.append("رفع سرعة الاستجابة وجودة التعامل في الخدمة عبر متابعة الأداء الميداني وتدريب الفريق.")
                elif aspect == 'الطعام':
                    recommendations.append("مراجعة جودة الأطباق وثبات التجربة بين الطلبات والفترات المختلفة.")
                elif aspect == 'السعر':
                    recommendations.append("إعادة تقييم القيمة المقدمة مقابل السعر أو توضيح المزايا التي تبرر التسعير الحالي.")
                elif aspect == 'النظافة':
                    recommendations.append("تشديد إجراءات النظافة والمتابعة الدورية داخل الفروع ونقاط التماس مع العميل.")
                elif aspect == 'التوصيل':
                    recommendations.append("تحسين زمن التوصيل وحالة الطلب عند الوصول وتقليل الأخطاء التشغيلية المرتبطة به.")
                elif aspect == 'الأجواء':
                    recommendations.append("تطوير الأجواء العامة داخل الفرع من حيث الراحة والتنظيم وتجربة الجلوس.")
                elif aspect == 'الموقع':
                    recommendations.append("معالجة الملاحظات المرتبطة بسهولة الوصول أو المواقف أو وضوح الموقع للعميل.")

        if not recommendations:
            recommendations.append("الاستمرار في مراقبة الجوانب الحالية مع التركيز على رفع نسبة التقييمات الإيجابية والمحايدة.")
            recommendations.append("تحليل التعليقات الجديدة بشكل دوري لاكتشاف أي تغير مبكر في تجربة العملاء.")

        return recommendations
